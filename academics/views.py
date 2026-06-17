# academics/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.db.models import Q, Avg, Max, Min, Count
from django.core.paginator import Paginator
import json
import re
from accounts.models import User, TeacherProfile
from .models import Subject, ClassLevel, AcademicYear, Term, ClassSubject, Result, SchoolDivision, ClassStream
import ast
from django.utils import timezone
from django.db import models

from datetime import datetime, date, timedelta
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.cell.cell import MergedCell
import openpyxl
import io


GES_BASIC_SUBJECTS = [
    {"name": "English Language", "code": "ENG", "category": "core"},
    {"name": "Mathematics", "code": "MATH", "category": "core"},
    {"name": "Science", "code": "SCI", "category": "core"},
    {"name": "Creative Arts and Design", "code": "CAD", "category": "core"},
    {"name": "Social Studies", "code": "SOC", "category": "core"},
    {"name": "Religious and Moral Education", "code": "RME", "category": "core"},
    {"name": "Computing", "code": "COMP", "category": "core"},
    {"name": "Career Technology", "code": "CTECH", "category": "core"},
    {"name": "Ghanaian Language", "code": "GHL", "category": "core"},
    {"name": "French", "code": "FR", "category": "elective"},
    {"name": "Arabic", "code": "AR", "category": "elective"},
    {"name": "Physical Education", "code": "PE", "category": "core"},
    {"name": "History", "code": "HIST", "category": "core"},
]


GES_BASIC_CLASS_LEVELS = [
    {"name": "Nursery 1", "code": "N1", "division": "Early Years"},
    {"name": "Nursery 2", "code": "N2", "division": "Early Years"},
    {"name": "Kindergarten 1", "code": "KG1", "division": "Kindergarten"},
    {"name": "Kindergarten 2", "code": "KG2", "division": "Kindergarten"},
    {"name": "Basic 1", "code": "B1", "division": "Primary"},
    {"name": "Basic 2", "code": "B2", "division": "Primary"},
    {"name": "Basic 3", "code": "B3", "division": "Primary"},
    {"name": "Basic 4", "code": "B4", "division": "Primary"},
    {"name": "Basic 5", "code": "B5", "division": "Primary"},
    {"name": "Basic 6", "code": "B6", "division": "Primary"},
    {"name": "JHS 1", "code": "JHS1", "division": "Junior High School"},
    {"name": "JHS 2", "code": "JHS2", "division": "Junior High School"},
    {"name": "JHS 3", "code": "JHS3", "division": "Junior High School"},
]


SCHOOL_CODE = "WSC"


def normalize_code(value):
    code = re.sub(r'[^A-Za-z0-9]+', '', value or '').upper()
    return code[:18] or "CODE"


def build_division_code(division_name, academic_year_badge):
    return f"{SCHOOL_CODE}-{normalize_code(division_name)}-{normalize_code(academic_year_badge)}"


def ensure_ges_basic_subjects():
    for item in GES_BASIC_SUBJECTS:
        Subject.objects.get_or_create(
            code=item["code"],
            defaults={
                "name": item["name"],
                "category": item["category"],
                "description": "GES/NaCCA basic school subject",
                "is_active": True,
            }
        )


def parse_academic_year_start(name):
    match = re.search(r'(\d{4})\D+(\d{2,4})', name or "")
    if not match:
        return None
    return int(match.group(1))


def ges_basic_calendar_for_year(start_year):
    next_year = start_year + 1
    return [
        {
            "name": "1st Term",
            "start_date": date(start_year, 9, 2),
            "end_date": date(start_year, 12, 18),
            "vacation_start_date": date(start_year, 12, 19),
            "vacation_end_date": date(next_year, 1, 7),
            "half_term_start_date": date(start_year, 10, 31),
            "half_term_end_date": date(start_year, 11, 3),
            "holidays": "Kwame Nkrumah Memorial Day; Farmers' Day",
            "school_activities": "Reopening; mid-term break; end-of-term vacation",
        },
        {
            "name": "2nd Term",
            "start_date": date(next_year, 1, 8),
            "end_date": date(next_year, 4, 1),
            "vacation_start_date": date(next_year, 4, 2),
            "vacation_end_date": date(next_year, 4, 20),
            "half_term_start_date": date(next_year, 2, 26),
            "half_term_end_date": date(next_year, 2, 27),
            "holidays": "Constitution Day; Independence Day; Eid ul-Fitr where applicable",
            "school_activities": "Reopening; mid-term break; vacation",
        },
        {
            "name": "3rd Term",
            "start_date": date(next_year, 4, 21),
            "end_date": date(next_year, 7, 23),
            "vacation_start_date": date(next_year, 7, 24),
            "vacation_end_date": None,
            "half_term_start_date": date(next_year, 6, 4),
            "half_term_end_date": date(next_year, 6, 5),
            "holidays": "May Day; Eid al-Adha where applicable; BECE period where applicable",
            "school_activities": "Reopening; mid-term break; end-of-year vacation",
        },
    ]


def ensure_ges_terms(academic_year):
    start_year = parse_academic_year_start(academic_year.name) or academic_year.start_date.year
    created = 0
    for index, term_data in enumerate(ges_basic_calendar_for_year(start_year), start=1):
        term, was_created = Term.objects.update_or_create(
            academic_year=academic_year,
            name=term_data["name"],
            defaults={
                **term_data,
                "is_current": index == 1 and academic_year.is_current,
            }
        )
        created += 1 if was_created else 0
    return created


def academic_year_terms_from_post(data, academic_year):
    start_year = parse_academic_year_start(academic_year.name) or academic_year.start_date.year
    defaults = ges_basic_calendar_for_year(start_year)
    terms = []

    for index, default in enumerate(defaults, start=1):
        prefix = f"term_{index}_"

        def date_value(field):
            value = data.get(f"{prefix}{field}")
            if value:
                return timezone.datetime.strptime(value, "%Y-%m-%d").date()
            return default.get(field)

        terms.append({
            "name": data.get(f"{prefix}name") or default["name"],
            "start_date": date_value("start_date"),
            "end_date": date_value("end_date"),
            "vacation_start_date": date_value("vacation_start_date"),
            "vacation_end_date": date_value("vacation_end_date"),
            "half_term_start_date": date_value("half_term_start_date"),
            "half_term_end_date": date_value("half_term_end_date"),
            "holidays": data.get(f"{prefix}holidays") or default.get("holidays"),
            "school_activities": data.get(f"{prefix}school_activities") or default.get("school_activities"),
            "is_current": data.get("current_term") == str(index),
        })

    return terms


def save_academic_year_terms(academic_year, term_data):
    for term in term_data:
        Term.objects.update_or_create(
            academic_year=academic_year,
            name=term["name"],
            defaults=term,
        )



@login_required
def subject_list(request):
    """List all subjects with filtering and pagination"""
    subjects = Subject.objects.all()
    
    category = request.GET.get('category')
    if category:
        subjects = subjects.filter(category=category)
    
    is_active = request.GET.get('is_active')
    if is_active:
        subjects = subjects.filter(is_active=is_active == 'true')
    
    search = request.GET.get('search')
    if search:
        subjects = subjects.filter(
            Q(name__icontains=search) |
            Q(code__icontains=search) |
            Q(description__icontains=search)
        )
    
    total_subjects = subjects.count()
    core_subjects = subjects.filter(category='core').count()
    elective_subjects = subjects.filter(category='elective').count()
    active_subjects = subjects.filter(is_active=True).count()
    
    paginator = Paginator(subjects, 9)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'category': category,
        'is_active': is_active,
        'search': search,
        'total_subjects': total_subjects,
        'core_subjects': core_subjects,
        'elective_subjects': elective_subjects,
        'active_subjects': active_subjects,
        'categories': Subject.SUBJECT_CATEGORY_CHOICES,
    }

    return render(request, 'pages/admin_dashboard/subject_lists.html', {'context': context})



@login_required
@require_http_methods(["POST"])
@transaction.atomic
def subject_create(request):
    """API endpoint for creating subjects"""
    try:
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
        
        name = data.get("name")
        code = data.get("code")
        description = data.get("description")
        category = data.get("category", "core")
        is_active = data.get("is_active") == "true" or data.get("is_active") == "on"
        
        if not all([name, code]):
            return JsonResponse({
                'success': False, 
                'error': 'Name and code are required.'
            }, status=400)
        
        if Subject.objects.filter(code=code).exists():
            return JsonResponse({
                'success': False, 
                'error': 'A subject with this code already exists.'
            }, status=400)
        
        subject = Subject.objects.create(
            name=name,
            code=code,
            description=description,
            category=category,
            is_active=is_active,
        )
        
        response_data = {
            'success': True,
            'message': f'Subject {name} created successfully.'
        }
        
        if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            messages.success(request, response_data['message'])
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while creating subject'
        }, status=400)


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def subject_update(request, subject_id):
    """API endpoint for updating subjects"""
    try:
        subject = get_object_or_404(Subject, id=subject_id)
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
        
        name = data.get("name")
        code = data.get("code")
        description = data.get("description")
        category = data.get("category")
        is_active = data.get("is_active") == "true" or data.get("is_active") == "on"
        
        if code != subject.code and Subject.objects.filter(code=code).exists():
            return JsonResponse({
                'success': False, 
                'error': 'A subject with this code already exists.'
            }, status=400)
        
        subject.name = name
        subject.code = code
        subject.description = description
        subject.category = category
        subject.is_active = is_active
        subject.save()
        
        response_data = {
            'success': True,
            'message': f'Subject {name} updated successfully.',
        }
        
        if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            messages.success(request, response_data['message'])
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while updating subject: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["DELETE", "POST"])
def subject_delete(request, subject_id):
    """Delete a subject"""
    try:
        subject = get_object_or_404(Subject, id=subject_id)
        subject_name = subject.name
        subject.delete()
        
        response_data = {
            'success': True,
            'message': f'Subject {subject_name} deleted successfully.',
        }
        
        if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            messages.success(request, response_data['message'])
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while deleting subject: {str(e)}'
        }, status=400)



@login_required
def class_list(request):
    """List all classes with filtering and pagination"""
    ensure_ges_basic_subjects()
    classes = ClassLevel.objects.prefetch_related('subjects').select_related('form_teacher').all()
    
    is_active = request.GET.get('is_active')
    if is_active:
        classes = classes.filter(is_active=is_active == 'true')
    
    search = request.GET.get('search')
    if search:
        classes = classes.filter(
            Q(name__icontains=search) |
            Q(code__icontains=search) |
            Q(division_name__icontains=search) |
            Q(division_code__icontains=search) |
            Q(description__icontains=search)
        )
    
    total_classes = classes.count()
    active_classes = classes.filter(is_active=True).count()
    total_capacity = sum(cls.capacity for cls in classes)
    total_students = sum(cls.current_students_count for cls in classes)
    
    paginator = Paginator(classes, 9)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'is_active': is_active,
        'search': search,
        'total_classes': total_classes,
        'active_classes': active_classes,
        'total_capacity': total_capacity,
        'total_students': total_students,
        'total_subjects': Subject.objects.count(),
        'ges_subjects': GES_BASIC_SUBJECTS,
        'ges_class_levels': GES_BASIC_CLASS_LEVELS,
        'current_academic_year': AcademicYear.objects.filter(is_current=True).first(),
        'subjects': Subject.objects.filter(is_active=True).order_by('name'),
        'teachers': User.objects.filter(role='teacher', is_active=True).order_by('first_name', 'last_name'),
    }
    return render(request, 'pages/admin_dashboard/class_list.html', {'context': context})


@login_required
@require_http_methods(["GET"])
def class_setup_options(request):
    academic_year = AcademicYear.objects.filter(is_current=True).first()
    requested_year = request.GET.get("academic_year") or (academic_year.name if academic_year else "")
    division_name = request.GET.get("division_name", "")

    return JsonResponse({
        "success": True,
        "school_code": SCHOOL_CODE,
        "academic_year_badge": normalize_code(requested_year),
        "division_code": build_division_code(division_name, requested_year) if division_name else "",
        "subjects": GES_BASIC_SUBJECTS,
        "class_levels": GES_BASIC_CLASS_LEVELS,
        "current_academic_year": {
            "id": academic_year.id,
            "name": academic_year.name,
        } if academic_year else None,
    })


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def class_create(request):
    """API endpoint for creating classes"""
    try:
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST

        if data.get("setup_mode") == "division_class_wizard":
            division_name = data.get("division_name", "").strip()
            division_code = data.get("division_code", "").strip()
            description = data.get("description", "").strip()
            academic_year_badge = data.get("academic_year_badge", "").strip()
            class_levels = data.get("class_levels", [])
            subject_items = data.get("subjects", [])
            academic_year = AcademicYear.objects.filter(is_current=True).first()

            if not division_name or not division_code or not academic_year_badge:
                return JsonResponse({
                    'success': False,
                    'error': 'Division name, division code, and academic year badge are required.'
                }, status=400)

            if not class_levels:
                return JsonResponse({
                    'success': False,
                    'error': 'Add at least one class level before saving.'
                }, status=400)

            division, _ = SchoolDivision.objects.update_or_create(
                code=division_code,
                defaults={
                    "name": division_name,
                    "description": description,
                    "is_active": True,
                }
            )

            subject_records = []
            for item in subject_items:
                subject_name = (item.get("name") or "").strip()
                subject_code = normalize_code(item.get("code") or subject_name)
                if not subject_name:
                    continue

                subject, _ = Subject.objects.get_or_create(
                    code=subject_code,
                    defaults={
                        "name": subject_name,
                        "description": item.get("description", ""),
                        "category": item.get("category", "core"),
                        "is_active": True,
                    }
                )
                if not subject.is_active:
                    subject.is_active = True
                    subject.save(update_fields=["is_active", "updated_at"])
                subject_records.append(subject)

            created_classes = []
            for index, item in enumerate(class_levels, start=1):
                level_name = (item.get("name") or "").strip()
                level_code = normalize_code(item.get("code") or level_name)
                capacity = int(item.get("capacity") or 30)
                if not level_name:
                    continue

                class_name = f"{level_name} - {division_name}"
                class_code = f"{division_code}-{level_code}"

                class_level, created = ClassLevel.objects.get_or_create(
                    name=class_name,
                    defaults={
                        "code": class_code,
                        "description": description,
                        "capacity": capacity,
                        "display_order": int(item.get("display_order") or index),
                        "is_active": True,
                        "division": division,
                        "division_name": division_name,
                        "division_code": division_code,
                        "academic_year_badge": academic_year_badge,
                    }
                )

                if not created:
                    class_level.code = class_code
                    class_level.description = description
                    class_level.capacity = capacity
                    class_level.display_order = int(item.get("display_order") or index)
                    class_level.is_active = True
                    class_level.division = division
                    class_level.division_name = division_name
                    class_level.division_code = division_code
                    class_level.academic_year_badge = academic_year_badge
                    class_level.save()

                if academic_year and subject_records:
                    for subject in subject_records:
                        ClassSubject.objects.get_or_create(
                            class_level=class_level,
                            subject=subject,
                            academic_year=academic_year,
                        )

                streams = item.get("streams") or ["A"]
                if isinstance(streams, str):
                    streams = [streams]
                for stream_name in streams:
                    stream_name = (stream_name or "A").strip().upper()
                    stream_code = f"{class_code}-{normalize_code(stream_name)}"
                    ClassStream.objects.update_or_create(
                        class_level=class_level,
                        name=stream_name,
                        defaults={
                            "code": stream_code,
                            "capacity": capacity,
                            "is_active": True,
                        }
                    )

                created_classes.append({
                    "id": str(class_level.id),
                    "name": class_level.name,
                    "code": class_level.code,
                    "created": created,
                })

            return JsonResponse({
                'success': True,
                'message': f'{len(created_classes)} class level(s) saved for {division_name}.',
                'classes': created_classes,
            })
        
        name = data.get("name")
        code = data.get("code")
        description = data.get("description")
        capacity = data.get("capacity", 30)
        display_order = data.get("display_order", 0)
        is_active = data.get("is_active") == "true" or data.get("is_active") == "on"
        form_teacher_id = data.get("form_teacher")
        division_name = data.get("division_name")
        division_code = data.get("division_code")
        academic_year_badge = data.get("academic_year_badge")
        subject_ids = data.getlist("subjects") if hasattr(data, 'getlist') else data.get("subjects", [])
        
        if not name:
            return JsonResponse({
                'success': False, 
                'error': 'Name is required.'
            }, status=400)
        
        if ClassLevel.objects.filter(name=name).exists():
            return JsonResponse({
                'success': False, 
                'error': 'A class with this name already exists.'
            }, status=400)
        
        if code and ClassLevel.objects.filter(code=code).exists():
            return JsonResponse({
                'success': False, 
                'error': 'A class with this code already exists.'
            }, status=400)
        
        class_level = ClassLevel.objects.create(
            name=name,
            code=code,
            description=description,
            capacity=capacity,
            display_order=display_order,
            is_active=is_active,
            form_teacher_id=form_teacher_id if form_teacher_id else None,
            division_name=division_name,
            division_code=division_code,
            academic_year_badge=academic_year_badge,
        )
        
        if subject_ids:
            if isinstance(subject_ids, str):
                import ast
                subject_ids = ast.literal_eval(subject_ids)
            subjects = Subject.objects.filter(id__in=subject_ids)
            academic_year = AcademicYear.objects.filter(is_current=True).first()
            if academic_year:
                for subject in subjects:
                    ClassSubject.objects.get_or_create(
                        class_level=class_level,
                        subject=subject,
                        academic_year=academic_year,
                    )
        
        response_data = {
            'success': True,
            'message': f'Class {name} created successfully.',
            'class_level': {
                'id': str(class_level.id),
                'name': class_level.name,
                'code': class_level.code,
                'capacity': class_level.capacity,
                'is_active': class_level.is_active,
            }
        }
        
        if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            messages.success(request, response_data['message'])
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while creating class: {str(e)}'
        }, status=400)



@login_required
@require_http_methods(["POST"])
@transaction.atomic
def class_update(request, class_id):
    """API endpoint for updating classes"""
    try:
        class_level = get_object_or_404(ClassLevel, id=class_id)
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
        
        name = data.get("name")
        code = data.get("code")
        description = data.get("description")
        capacity = data.get("capacity")
        display_order = data.get("display_order")
        is_active = data.get("is_active") == "true" or data.get("is_active") == "on"
        form_teacher_id = data.get("form_teacher")
        subject_ids = data.getlist("subjects") if hasattr(data, 'getlist') else data.get("subjects", [])
        
        if name != class_level.name and ClassLevel.objects.filter(name=name).exists():
            return JsonResponse({
                'success': False, 
                'error': 'A class with this name already exists.'
            }, status=400)
        
        if code != class_level.code and ClassLevel.objects.filter(code=code).exists():
            return JsonResponse({
                'success': False, 
                'error': 'A class with this code already exists.'
            }, status=400)
        
        class_level.name = name
        class_level.code = code
        class_level.description = description
        class_level.capacity = capacity
        class_level.display_order = display_order
        class_level.is_active = is_active
        class_level.form_teacher_id = form_teacher_id if form_teacher_id else None
        class_level.save()
        
        if subject_ids:
            if isinstance(subject_ids, str):
                import ast
                subject_ids = ast.literal_eval(subject_ids)
            subjects = Subject.objects.filter(id__in=subject_ids)
            class_level.subjects.set(subjects)
        
        response_data = {
            'success': True,
            'message': f'Class {name} updated successfully.',
        }
        
        if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            messages.success(request, response_data['message'])
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while updating class: {str(e)}'
        }, status=400)



@login_required
@require_http_methods(["DELETE", "POST"])
def class_delete(request, class_id):
    """Delete a class"""
    try:
        class_level = get_object_or_404(ClassLevel, id=class_id)
        class_name = class_level.name
        
        # Check if class has students
        if class_level.students.exists():
            return JsonResponse({
                'success': False, 
                'error': f'Cannot delete {class_name} because it has students assigned. Please reassign students first.'
            }, status=400)
        
        class_level.delete()
        
        response_data = {
            'success': True,
            'message': f'Class {class_name} deleted successfully.',
        }
        
        if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            messages.success(request, response_data['message'])
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while deleting class: {str(e)}'
        }, status=400)



@login_required
@require_http_methods(["GET"])
def get_subject_data(request, subject_id):
    """Get subject data for AJAX requests"""
    try:
        subject = get_object_or_404(Subject, id=subject_id)
        
        data = {
            'id': str(subject.id),
            'name': subject.name,
            'code': subject.code,
            'description': subject.description,
            'category': subject.category,
            'is_active': subject.is_active,
            'created_at': subject.created_at.isoformat(),
            'updated_at': subject.updated_at.isoformat(),
        }
        
        return JsonResponse({'success': True, 'data': data})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def get_class_data(request, class_id):
    """Get class data for AJAX requests"""
    try:
        class_level = get_object_or_404(ClassLevel, id=class_id)
        subject_handlers = []
        assignments = ClassSubject.objects.filter(class_level=class_level).select_related('subject', 'teacher', 'academic_year')
        for assignment in assignments:
            subject_handlers.append({
                'subject_id': str(assignment.subject.id),
                'subject_name': assignment.subject.name,
                'subject_code': assignment.subject.code,
                'teacher_id': str(assignment.teacher.id) if assignment.teacher else None,
                'teacher_name': assignment.teacher.get_full_name() if assignment.teacher else 'Not assigned',
                'academic_year': assignment.academic_year.name if assignment.academic_year else '',
            })
        
        data = {
            'id': str(class_level.id),
            'name': class_level.name,
            'code': class_level.code,
            'division_name': class_level.division_name,
            'division_code': class_level.division_code,
            'academic_year_badge': class_level.academic_year_badge,
            'description': class_level.description,
            'capacity': class_level.capacity,
            'display_order': class_level.display_order,
            'is_active': class_level.is_active,
            'form_teacher': str(class_level.form_teacher.id) if class_level.form_teacher else None,
            'form_teacher_name': class_level.form_teacher.get_full_name() if class_level.form_teacher else None,
            'current_students_count': class_level.current_students_count,
            'available_seats': class_level.available_seats,
            'subjects': list(class_level.subjects.values('id', 'name', 'code', 'category')),
            'subject_handlers': subject_handlers,
            'streams': list(class_level.streams.values('id', 'name', 'code', 'capacity', 'is_active')),
            'created_at': class_level.created_at.isoformat(),
            'updated_at': class_level.updated_at.isoformat(),
        }
        
        return JsonResponse({'success': True, 'data': data})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def get_teachers_list(request):
    """Get list of teachers for dropdowns"""
    try:
        teachers = User.objects.filter(
            role='teacher',
            teacher_profile__is_active=True
        ).select_related('teacher_profile').values(
            'id', 'first_name', 'last_name', 'email', 'teacher_profile__employee_id'
        )
        
        teachers_list = list(teachers)
        return JsonResponse({'success': True, 'teachers': teachers_list})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def get_subjects_list(request):
    """Get list of subjects for dropdowns"""
    try:
        subjects = Subject.objects.filter(is_active=True).values('id', 'name', 'code', 'category')
        subjects_list = list(subjects)
        return JsonResponse({'success': True, 'subjects': subjects_list})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)



@login_required
@require_http_methods(["GET", "POST"])
def create_academic_year(request):
    """Create a new academic year"""
    if request.method == 'POST':
        try:
            data = request.POST
            
            name = data.get('name')
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            is_current = data.get('is_current') == 'on'
            
            if name and (not start_date or not end_date):
                start_year = parse_academic_year_start(name)
                if start_year:
                    generated_terms = ges_basic_calendar_for_year(start_year)
                    start_date = generated_terms[0]["start_date"].isoformat()
                    end_date = generated_terms[-1]["end_date"].isoformat()

            if not all([name, start_date, end_date]):
                return JsonResponse({
                    'success': False,
                    'error': 'Academic year name is required. Use a format like 2025/2026 so dates can be generated.'
                }, status=400)
            

            start_date_obj = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
            
            if start_date_obj >= end_date_obj:
                return JsonResponse({
                    'success': False,
                    'error': 'End date must be after start date'
                }, status=400)
            
            overlapping_years = AcademicYear.objects.filter(
                models.Q(start_date__lte=end_date_obj) & models.Q(end_date__gte=start_date_obj)
            )
            
            if overlapping_years.exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Academic year overlaps with existing academic year'
                }, status=400)


            academic_year = AcademicYear(
                name=name,
                start_date=start_date_obj,
                end_date=end_date_obj,
                is_current=is_current
            )
            
            academic_year.save()
            term_data = academic_year_terms_from_post(data, academic_year)
            if not any(term["is_current"] for term in term_data) and academic_year.is_current and term_data:
                term_data[0]["is_current"] = True
            save_academic_year_terms(academic_year, term_data)
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Academic year "{name}" created successfully',
                    'academic_year': {
                        'id': str(academic_year.id),
                        'name': academic_year.name,
                        'start_date': academic_year.start_date.isoformat(),
                        'end_date': academic_year.end_date.isoformat(),
                        'is_current': academic_year.is_current
                    }
                })
            else:
                messages.success(request, f'Academic year "{name}" created successfully')
                return redirect('academic_years_list')
                
        except ValueError as e:
            return JsonResponse({
                'success': False,
                'error': 'Invalid date format. Use YYYY-MM-DD'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error creating academic year: {str(e)}'
            }, status=400)
    
    return render(request, 'pages/academics/create_academic_year.html')


@login_required
def get_academic_years(request):
    """API endpoint to get all academic years for dropdown"""
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    
    years_data = [
        {
            'id': str(year.id),
            'name': year.name,
            'is_current': year.is_current,
            'start_date': year.start_date.isoformat(),
            'end_date': year.end_date.isoformat()
        }
        for year in academic_years
    ]
    
    return JsonResponse({'academic_years': years_data})


@login_required
def academic_years_list(request):
    """List all academic years"""
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    return render(request, 'pages/academics/academic_years_list.html', {
        'academic_years': academic_years
    })


@require_http_methods(["PUT"])
def edit_academic_year(request, year_id):
    """Edit an academic year"""
    try:
        academic_year = get_object_or_404(AcademicYear, id=year_id)
        data = json.loads(request.body)
        
        academic_year.name = data.get('name', academic_year.name)
        academic_year.start_date = data.get('start_date', academic_year.start_date)
        academic_year.end_date = data.get('end_date', academic_year.end_date)
        academic_year.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Academic year updated successfully',
            'academic_year': {
                'id': academic_year.id,
                'name': academic_year.name,
                'start_date': academic_year.start_date,
                'end_date': academic_year.end_date,
                'is_current': academic_year.is_current
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@require_http_methods(["POST"])
def set_current_academic_year(request, year_id):
    """Set an academic year as current"""
    try:
        academic_year = get_object_or_404(AcademicYear, id=year_id)
        academic_year.is_current = True
        academic_year.save()
        
        return JsonResponse({
            'success': True,
            'message': f'{academic_year.name} set as current academic year'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@require_http_methods(["DELETE"])
def delete_academic_year(request, year_id):
    """Delete an academic year"""
    try:
        academic_year = get_object_or_404(AcademicYear, id=year_id)
        
        if academic_year.is_current:
            return JsonResponse({
                'success': False,
                'error': 'Cannot delete the current academic year'
            }, status=400)
        
        academic_year_name = academic_year.name
        academic_year.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Academic year {academic_year_name} deleted successfully'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
    



@login_required
@require_http_methods(["POST"])
@transaction.atomic
def assign_subjects_to_class(request, class_id):
    """API endpoint for assigning subjects to a class"""
    try:
        class_level = get_object_or_404(ClassLevel, id=class_id)
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
        
        subject_ids = data.getlist("subjects") if hasattr(data, 'getlist') else data.get("subjects", [])
        academic_year_id = data.get("academic_year")
        
        if isinstance(subject_ids, str):
            subject_ids = ast.literal_eval(subject_ids)

        try:
            academic_year = AcademicYear.objects.get(id=academic_year_id)
        except AcademicYear.DoesNotExist:
            academic_year = AcademicYear.objects.filter(is_current=True).first()
            if not academic_year:
                return JsonResponse({
                    'success': False,
                    'error': 'Create or select an academic year before assigning subjects.'
                }, status=400)
        
        subjects = Subject.objects.filter(id__in=subject_ids)
        
        ClassSubject.objects.filter(
            class_level=class_level, 
            academic_year=academic_year
        ).delete()
        
        class_subjects = []
        for subject in subjects:
            class_subjects.append(
                ClassSubject(
                    class_level=class_level,
                    subject=subject,
                    academic_year=academic_year
                )
            )
        
        ClassSubject.objects.bulk_create(class_subjects)
        
        response_data = {
            'success': True,
            'message': f'Successfully updated subjects for {class_level.name}.',
            'subjects_count': subjects.count(),
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while updating class subjects: {str(e)}'
        }, status=400)
    

@login_required
@require_http_methods(["POST"])
@transaction.atomic
def assign_teacher_to_class(request, class_id):
    """API endpoint for assigning teacher to a class"""
    try:
        class_level = get_object_or_404(ClassLevel, id=class_id)
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
        
        teacher_id = data.get("form_teacher")
        
        class_level.form_teacher_id = teacher_id if teacher_id else None
        class_level.save()
        
        teacher_name = class_level.form_teacher.get_full_name() if class_level.form_teacher else "No teacher"
        
        response_data = {
            'success': True,
            'message': f'Successfully assigned {teacher_name} to {class_level.name}.',
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while assigning teacher: {str(e)}'
        }, status=400)
    

    
# @login_required
# @require_http_methods(["POST"])
# def assign_teacher_to_subject(request, subject_id):
#     """API endpoint for assigning teacher to a subject"""
#     try:
#         subject = get_object_or_404(Subject, id=subject_id)
        
#         teacher_id = request.POST.get("teacher_id")
        
#         if teacher_id:
#             teacher = get_object_or_404(User, id=teacher_id, role='teacher')
#             subject.teacher = teacher
#         else:
#             subject.teacher = None
            
#         subject.save()
        
#         teacher_name = subject.teacher.get_full_name() if subject.teacher else "No teacher"
        
#         response_data = {
#             'success': True,
#             'message': f'Successfully assigned {teacher_name} to {subject.name}.',
#         }
        
#         return JsonResponse(response_data)
        
#     except Exception as e:
#         return JsonResponse({
#             'success': False, 
#             'error': f'An error occurred while assigning teacher: {str(e)}'
#         }, status=400)


@login_required
def get_teachers_api(request):
    """API endpoint to get all teachers for dropdown"""
    teachers = User.objects.filter(role='teacher', is_active=True).order_by('first_name')
    
    teachers_data = [
        {
            'id': str(teacher.id),
            'name': teacher.get_full_name(),
            'email': teacher.email
        }
        for teacher in teachers
    ]
    
    return JsonResponse({
        'success': True,
        'teachers': teachers_data
    })



@login_required
def manage_subject_teachers(request, subject_id):
    """View to manage teachers for a subject across different classes"""
    subject = get_object_or_404(Subject, id=subject_id)
    
    class_assignments = ClassSubject.objects.filter(
        subject=subject
    ).select_related('class_level', 'teacher', 'academic_year')
    
    classes = ClassLevel.objects.filter(is_active=True)
    teachers = User.objects.filter(role='teacher', is_active=True)
    academic_years = AcademicYear.objects.all()
    
    context = {
        'subject': subject,
        'class_assignments': class_assignments,
        'classes': classes,
        'teachers': teachers,
        'academic_years': academic_years,
    }
    
    return render(request, 'pages/admin_dashboard/manage_subject_teachers.html', context)



@login_required
@require_http_methods(["POST"])
def assign_teacher_to_class_subject(request):
    """API endpoint for assigning teacher to a subject for a specific class"""
    try:
        subject_id = request.POST.get("subject_id")
        class_level_id = request.POST.get("class_level_id")
        teacher_id = request.POST.get("teacher_id")
        academic_year_id = request.POST.get("academic_year_id")
        
        subject = get_object_or_404(Subject, id=subject_id)
        class_level = get_object_or_404(ClassLevel, id=class_level_id)
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        teacher = get_object_or_404(User, id=teacher_id, role='teacher') if teacher_id else None
        
        # Check if assignment already exists
        existing_assignment = ClassSubject.objects.filter(
            class_level=class_level,
            subject=subject,
            academic_year=academic_year
        ).first()
        
        if existing_assignment:
            existing_assignment.teacher = teacher
            existing_assignment.save()
            action = "updated"
        else:
            ClassSubject.objects.create(
                class_level=class_level,
                subject=subject,
                academic_year=academic_year,
                teacher=teacher
            )
            action = "added"
        
        teacher_name = teacher.get_full_name() if teacher else "No teacher"
        
        response_data = {
            'success': True,
            'message': f'Successfully {action} {teacher_name} to {subject.name} for {class_level.name}.',
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while assigning teacher: {str(e)}'
        }, status=400)



@login_required
@require_http_methods(["POST"])
def update_class_assignment(request):
    """API endpoint for updating class assignment teacher"""
    try:
        assignment_id = request.POST.get("assignment_id")
        teacher_id = request.POST.get("teacher_id")
        
        assignment = get_object_or_404(ClassSubject, id=assignment_id)
        teacher = get_object_or_404(User, id=teacher_id, role='teacher') if teacher_id else None
        
        assignment.teacher = teacher
        assignment.save()
        
        teacher_name = teacher.get_full_name() if teacher else "No teacher"
        
        response_data = {
            'success': True,
            'message': f'Successfully updated teacher to {teacher_name} for {assignment.class_level.name} - {assignment.subject.name}.',
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while updating assignment: {str(e)}'
        }, status=400)



@login_required
@require_http_methods(["DELETE"])
def delete_class_assignment(request, assignment_id):
    """API endpoint for deleting a class assignment"""
    try:
        assignment = get_object_or_404(ClassSubject, id=assignment_id)
        class_name = assignment.class_level.name
        subject_name = assignment.subject.name
        
        assignment.delete()
        
        response_data = {
            'success': True,
            'message': f'Successfully removed assignment for {class_name} - {subject_name}.',
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while removing assignment: {str(e)}'
        }, status=400)  
    



@login_required
def results_dashboard(request):
    """Main results dashboard for teachers and admin"""
    
    # Get filter parameters
    academic_year_id = request.GET.get('academic_year')
    class_level_id = request.GET.get('class_level')
    term_id = request.GET.get('term')
    subject_id = request.GET.get('subject')
    
    # Base queryset
    if request.user.role == 'teacher':
        results = Result.objects.filter(
            Q(uploaded_by=request.user) | Q(subject__teacher=request.user)
        )
    else:
        results = Result.objects.all()
    
    # Apply filters
    if academic_year_id:
        results = results.filter(term__academic_year_id=academic_year_id)
    if class_level_id:
        results = results.filter(class_level_id=class_level_id)
    if term_id:
        results = results.filter(term_id=term_id)
    if subject_id:
        results = results.filter(subject_id=subject_id)
    
    # Get available filters
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    class_levels = ClassLevel.objects.filter(is_active=True)
    terms = Term.objects.all().order_by('-academic_year', 'start_date')
    
    if request.user.role == 'teacher':
        subjects = Subject.objects.filter(
            classsubject__teacher=request.user
        ).distinct()
    else:
        subjects = Subject.objects.all()

    
    # Statistics
    total_results = results.count()
    published_results = results.filter(is_published=True).count()
    average_score = results.aggregate(avg=Avg('score'))['avg'] or 0
    
    # Recent results
    recent_results = results.select_related(
        'student', 'subject', 'class_level', 'term'
    ).order_by('-date_uploaded')[:10]
    
    # Performance by subject
    subject_performance = results.values(
        'subject__name', 'subject__code'
    ).annotate(
        avg_score=Avg('score'),
        count=Count('id'),
        max_score=Max('score'),
        min_score=Min('score')
    ).order_by('-avg_score')
    
    context = {
        'total_results': total_results,
        'published_results': published_results,
        'average_score': round(average_score, 2),
        'recent_results': recent_results,
        'subject_performance': subject_performance,
        'academic_years': academic_years,
        'class_levels': class_levels,
        'terms': terms,
        'subjects': subjects,
        'current_filters': {
            'academic_year_id': academic_year_id,
            'class_level_id': class_level_id,
            'term_id': term_id,
            'subject_id': subject_id,
        }
    }
    
    return render(request, 'pages/admin_dashboard/results.html', {'context': context})


@login_required
def upload_results_form(request):
    """GET: Show the upload results form"""

    if request.user.role != 'teacher':
        return render(request, 'pages/academics/results/not_allowed.html', {
            'message': 'Only teachers can upload results.'
        }, status=403)

    # Only classes this teacher actually teaches
    class_levels = ClassLevel.objects.filter(
        classsubject__teacher=request.user,
        is_active=True
    ).distinct()

    # Only subjects this teacher teaches
    subjects = Subject.objects.filter(
        classsubject__teacher=request.user
    ).distinct()

    academic_years = AcademicYear.objects.all().order_by('-start_date')

    context = {
        'academic_years': academic_years,
        'class_levels': class_levels,
        'subjects': subjects,
    }

    return render(request, 'pages/academics/results/upload_result.html', context)



@login_required
def upload_results_submit(request):
    """Process ONE student's result"""

    if request.user.role != 'teacher':
        return JsonResponse({'success': False, 'error': "Only teachers can upload results."}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': "Invalid request method."}, status=400)

    try:
        data = request.POST

        student_id = data.get("student")
        class_level_id = data.get("class_level")
        subject_id = data.get("subject")
        term_id = data.get("term")
        academic_year_id = data.get("academic_year")

        calculation_mode = data.get("calculation_mode", "system")  # system | manual
        class_score = data.get("class_score")
        exam_score = data.get("exam_score")
        manual_score = data.get("score")

        is_published = data.get("is_published") == "on"
        remarks = data.get("remarks", "").strip()

        required = {
            'student': student_id,
            'class_level': class_level_id,
            'subject': subject_id,
            'term': term_id,
            'academic_year': academic_year_id
        }

        missing = [f for f, v in required.items() if not v]
        if missing:
            return JsonResponse({
                'success': False,
                'error': f"Missing required fields: {', '.join(missing)}"
            }, status=400)

        student = get_object_or_404(User, id=student_id, role="student")
        class_level = get_object_or_404(ClassLevel, id=class_level_id)
        subject = get_object_or_404(Subject, id=subject_id)
        term = get_object_or_404(Term, id=term_id)

        if student.student_profile.current_class != class_level:
            return JsonResponse({
                'success': False,
                'error': f"Student {student.get_full_name()} is not in class {class_level.name}."
            }, status=400)

        if not ClassSubject.objects.filter(
            class_level=class_level,
            subject=subject,
            teacher=request.user
        ).exists():
            return JsonResponse({
                'success': False,
                'error': "You are not assigned to teach this subject for this class."
            }, status=403)

        score_data = {}

        if calculation_mode == "system":
            try:
                class_val = float(class_score or 0)
                exam_val = float(exam_score or 0)
            except:
                return JsonResponse({'success': False, 'error': "Invalid score format."}, status=400)

            if not (0 <= class_val <= 100):
                return JsonResponse({'success': False, 'error': "Class score must be 0-100."}, status=400)

            if not (0 <= exam_val <= 100):
                return JsonResponse({'success': False, 'error': "Exam score must be 0-100."}, status=400)

            score_data = {
                'class_score': class_val,
                'exam_score': exam_val,
                'score': None,
            }

        else:
            try:
                manual_val = float(manual_score or 0)
                class_val = float(class_score or 0)
                exam_val = float(exam_score or 0)
            except:
                return JsonResponse({
                    'success': False,
                    'error': "Invalid numeric input."
                }, status=400)

            if not (0 <= manual_val <= 100):
                return JsonResponse({
                    'success': False,
                    'error': "Total score must be 0-100."
                }, status=400)

            score_data = {
                'class_score': class_val,
                'exam_score': exam_val,
                'score': manual_val,
            }

        result_data = {
            **score_data,
            'class_level': class_level,
            'term': term,
            'uploaded_by': request.user,
            'remarks': remarks if remarks else None,
            'is_published': is_published,
        }

        if is_published:
            result_data['published_date'] = timezone.now()

        result, created = Result.objects.update_or_create(
            student=student,
            subject=subject,
            term=term,
            defaults=result_data
        )

        return JsonResponse({
            'success': True,
            'created': created,
            'message': f"Result {'created' if created else 'updated'} successfully."
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f"Unexpected error: {str(e)}"
        }, status=500)


@login_required
def get_students_for_results(request):
    """Get students for a specific class to populate results form"""
    class_level_id = request.GET.get('class_level_id')
    academic_year_id = request.GET.get('academic_year_id')
    
    if not class_level_id:
        return JsonResponse({'success': False, 'error': 'Class level ID required'})
    
    class_level = get_object_or_404(ClassLevel, id=class_level_id)
    
    students = User.objects.filter(
        role='student',
        student_profile__current_class=class_level,
        student_profile__is_active=True
    ).select_related('student_profile').values(
        'id',
        'first_name',
        'last_name',
        'student_profile__student_id'
    ).order_by('first_name', 'last_name')
    
    return JsonResponse({
        'success': True,
        'students': list(students)
    })


@login_required
def get_existing_results(request):
    """Get existing results for a class, subject, and term"""
    class_level_id = request.GET.get('class_level_id')
    subject_id = request.GET.get('subject_id')
    term_id = request.GET.get('term_id')
    
    if not all([class_level_id, subject_id, term_id]):
        return JsonResponse({'success': False, 'error': 'All parameters required'})
    
    results = Result.objects.filter(
        class_level_id=class_level_id,
        subject_id=subject_id,
        term_id=term_id
    ).select_related('student').values(
        'student_id',
        'score',
        'class_score',
        'exam_score',
        'grade',
        'is_published'
    )
    
    results_dict = {str(result['student_id']): result for result in results}

    
    return JsonResponse({
        'success': True,
        'results': results_dict
    })


@login_required
@require_http_methods(["POST"])
def publish_results(request, result_id=None):
    """Publish or unpublish results"""
    try:
        if result_id:
            # Publish single result
            result = get_object_or_404(Result, id=result_id)
            
            # Check permissions
            if request.user.role == 'teacher' and result.uploaded_by != request.user:
                return JsonResponse({
                    'success': False,
                    'error': 'You can only publish results you uploaded'
                }, status=403)
            
            result.is_published = not result.is_published
            result.save()
            
            action = 'published' if result.is_published else 'unpublished'
            
            return JsonResponse({
                'success': True,
                'message': f'Result {action} successfully',
                'is_published': result.is_published
            })
        else:
            # Bulk publish
            data = json.loads(request.body)
            result_ids = data.get('result_ids', [])
            publish = data.get('publish', True)
            
            results = Result.objects.filter(id__in=result_ids)
            
            # Check permissions for teachers
            if request.user.role == 'teacher':
                results = results.filter(uploaded_by=request.user)
            
            updated_count = results.update(is_published=publish)
            
            action = 'published' if publish else 'unpublished'
            
            return JsonResponse({
                'success': True,
                'message': f'{updated_count} results {action} successfully'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error publishing results: {str(e)}'
        }, status=400)


@login_required
def result_detail(request, result_id):
    """Detailed view of a single result"""

    user = request.user
    
     # Determine which base template to use based on role

    if user.role == 'admin':
        base_template = 'base/admin_d.html'
    elif user.role == 'teacher':
        base_template = 'base/teacher_d.html'
    elif user.role == 'student':
        base_template = 'base/student_d.html'
    else:
        base_template = 'base/admin_d.html'
            
    result = get_object_or_404(
        Result.objects.select_related(
            'student', 'subject', 'class_level', 'term', 'uploaded_by'
        ),
        id=result_id
    )
    
    # Check permissions
    if request.user.role == 'student' and result.student != request.user:
        messages.error(request, 'You can only view your own results')
        return redirect('student_results')
    
    if request.user.role == 'teacher' and result.uploaded_by != request.user:
        messages.error(request, 'You can only view results you uploaded')
        return redirect('results_dashboard')
    
    context = {
        'result': result,
        'base_template': base_template
    }
    
    return render(request, 'pages/academics/results/detail.html', {'context': context})


@login_required
@require_http_methods(["DELETE"])
def delete_result(request, result_id):
    """Delete a result"""
    try:
        result = get_object_or_404(Result, id=result_id)
        
        if request.user.role == 'teacher' and result.uploaded_by != request.user:
            return JsonResponse({
                'success': False,
                'error': 'You can only delete results you uploaded'
            }, status=403)
        
        result.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Result deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error deleting result: {str(e)}'
        }, status=400)



@login_required
def results_analysis(request):
    """Advanced results analysis"""

    academic_year_id = request.GET.get('academic_year')
    class_level_id = request.GET.get('class_level')
    term_id = request.GET.get('term')
    
    results = Result.objects.select_related(
        'student', 'subject', 'class_level', 'term', 'term__academic_year'
    ).filter(is_published=True)
    
    if academic_year_id:
        results = results.filter(term__academic_year_id=academic_year_id)
    if class_level_id:
        results = results.filter(class_level_id=class_level_id)
    if term_id:
        results = results.filter(term_id=term_id)
    
    # Comprehensive analysis
    grade_distribution = results.values('grade').annotate(
        count=Count('id')
    ).order_by('grade')
    
    # Subject performance (average scores per subject)
    subject_performance = results.values('subject__name').annotate(
        avg_score=Avg('score')
    ).order_by('-avg_score')
    
    class_performance = results.values('class_level__name').annotate(
        avg_score=Avg('score'),
        total_students=Count('student', distinct=True),
        pass_rate=Count('id', filter=Q(score__gte=50)) * 100.0 / Count('id')
    ).order_by('-avg_score')
    
    performance_trends = results.values(
        'subject__name',
        'term__name',
        'term__academic_year__name'
    ).annotate(
        avg_score=Avg('score'),
        student_count=Count('student', distinct=True)
    ).order_by('term__academic_year__name', 'term__name', 'subject__name')
    
    top_performers = results.select_related('student', 'subject').order_by('-score')[:10]
    
    # Convert Decimal to float for JSON serialization
    from decimal import Decimal
    
    subject_performance_list = []
    for item in subject_performance:
        subject_performance_list.append({
            'subject__name': item['subject__name'],
            'avg_score': float(item['avg_score']) if isinstance(item['avg_score'], Decimal) else item['avg_score']
        })
    
    class_performance_list = []
    for item in class_performance:
        class_performance_list.append({
            'class_level__name': item['class_level__name'],
            'avg_score': float(item['avg_score']) if isinstance(item['avg_score'], Decimal) else item['avg_score'],
            'total_students': item['total_students'],
            'pass_rate': float(item['pass_rate']) if isinstance(item['pass_rate'], Decimal) else item['pass_rate']
        })
    
    # Convert performance_trends
    performance_trends_list = []
    for item in performance_trends:
        performance_trends_list.append({
            'subject__name': item['subject__name'],
            'term__name': item['term__name'],
            'term__academic_year__name': item['term__academic_year__name'],
            'avg_score': float(item['avg_score']) if isinstance(item['avg_score'], Decimal) else item['avg_score'],
            'student_count': item['student_count']
        })
    
    academic_years = AcademicYear.objects.all()
    class_levels = ClassLevel.objects.all()
    terms = Term.objects.all()
    
    context = {
        'grade_distribution': list(grade_distribution),
        'subject_performance': subject_performance_list,
        'class_performance': class_performance_list,
        'performance_trends': performance_trends_list,
        'top_performers': top_performers,
        'total_results_analyzed': results.count(),
        'academic_years': academic_years,
        'class_levels': class_levels,
        'terms': terms,
        'current_filters': {
            'academic_year_id': academic_year_id,
            'class_level_id': class_level_id,
            'term_id': term_id,
        }
    }
    
    return render(request, 'pages/academics/results/analysis.html', {'context': context})


@login_required
def get_terms_for_academic_year(request):
    """Get terms for a specific academic year"""
    academic_year_id = request.GET.get('academic_year')
    
    # if not academic_year_id:
    #     return JsonResponse({'success': False, 'error': 'Academic year ID required'})

    # NOTE:
    # We are NOT validating academic_year_id here.
    # The <select> on the frontend loads with an empty default option (no value).
    # If we enforce "academic_year_id is required", the client would receive an 
    # unnecessary error before the teacher actually selects a year.
    # Therefore, we skip the check to avoid sending an error when the field is still empty.

    
    terms = Term.objects.filter(academic_year_id=academic_year_id).values(
        'id', 'name', 'start_date', 'end_date'
    ).order_by('start_date')
    
    return JsonResponse({
        'success': True,
        'terms': list(terms)
    })





@login_required
def export_analysis_report(request):
    """Export analysis report as PDF or Excel"""
    format_type = request.GET.get('format', 'pdf')
    academic_year_id = request.GET.get('academic_year')
    class_level_id = request.GET.get('class_level')
    term_id = request.GET.get('term')
    
    # Get filtered data
    analysis_data = get_analysis_data(academic_year_id, class_level_id, term_id)
    
    if format_type.lower() == 'excel':
        return export_excel_report(analysis_data, request)
    else:
        return export_pdf_report(analysis_data, request)

def get_analysis_data(academic_year_id=None, class_level_id=None, term_id=None):
    """Get comprehensive analysis data for export"""
    results = Result.objects.all()
    
    # Apply filters
    if academic_year_id:
        results = results.filter(term__academic_year_id=academic_year_id)
    if class_level_id:
        results = results.filter(class_level_id=class_level_id)
    if term_id:
        results = results.filter(term_id=term_id)
    
    # Comprehensive analysis data
    data = {
        'filters': {
            'academic_year': AcademicYear.objects.filter(id=academic_year_id).first() if academic_year_id else None,
            'class_level': ClassLevel.objects.filter(id=class_level_id).first() if class_level_id else None,
            'term': Term.objects.filter(id=term_id).first() if term_id else None,
        },
        'summary': {
            'total_results': results.count(),
            'total_students': results.values('student').distinct().count(),
            'average_score': results.aggregate(avg=Avg('score'))['avg'] or 0,
            'published_results': results.filter(is_published=True).count(),
        },
        'grade_distribution': list(results.values('grade').annotate(
            count=Count('id')
        ).order_by('grade')),
        'subject_performance': list(results.values(
            'subject__name', 'subject__code'
        ).annotate(
            avg_score=Avg('score'),
            total_students=Count('student', distinct=True),
            pass_count=Count('id', filter=Q(score__gte=50)),
            max_score=Max('score'),
            min_score=Min('score')
        ).order_by('-avg_score')),
        'class_performance': list(results.values('class_level__name').annotate(
            avg_score=Avg('score'),
            total_students=Count('student', distinct=True),
            pass_rate=Count('id', filter=Q(score__gte=50)) * 100.0 / Count('id')
        ).order_by('-avg_score')),
        'top_performers': list(results.select_related(
            'student', 'subject', 'class_level'
        ).order_by('-score')[:10].values(
            'student__first_name',
            'student__last_name',
            'student__student_profile__student_id',
            'subject__name',
            'class_level__name',
            'score',
            'grade'
        )),
        'generated_at': datetime.now()
    }
    
    return data

def export_pdf_report(analysis_data, request):
    """Generate PDF report"""
    try:
        # Create buffer for PDF
        buffer = io.BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )
        
        # Get styles
        styles = getSampleStyleSheet()
        
        # Create custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1,  # Center
            textColor=colors.HexColor('#1e293b')
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            spaceBefore=20,
            textColor=colors.HexColor('#374151')
        )
        
        # Build story (content)
        story = []
        
        # Title
        title = Paragraph("Academic Results Analysis Report", title_style)
        story.append(title)
        
        # Report metadata
        metadata_style = ParagraphStyle(
            'Metadata',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#64748b')
        )
        
        filters_text = "Filters: "
        filters = []
        if analysis_data['filters']['academic_year']:
            filters.append(f"Academic Year: {analysis_data['filters']['academic_year'].name}")
        if analysis_data['filters']['class_level']:
            filters.append(f"Class: {analysis_data['filters']['class_level'].name}")
        if analysis_data['filters']['term']:
            filters.append(f"Term: {analysis_data['filters']['term'].name}")
        
        filters_text += " | ".join(filters) if filters else "All Data"
        filters_text += f" | Generated: {analysis_data['generated_at'].strftime('%Y-%m-%d %H:%M')}"
        
        metadata = Paragraph(filters_text, metadata_style)
        story.append(metadata)
        story.append(Spacer(1, 20))
        
        # Summary Section
        story.append(Paragraph("Executive Summary", heading_style))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Results', f"{analysis_data['summary']['total_results']:,}"],
            ['Total Students', f"{analysis_data['summary']['total_students']:,}"],
            ['Average Score', f"{analysis_data['summary']['average_score']:.1f}%"],
            ['Published Results', f"{analysis_data['summary']['published_results']:,}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0'))
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Grade Distribution
        story.append(Paragraph("Grade Distribution", heading_style))
        
        grade_data = [['Grade', 'Count', 'Percentage']]
        total_results = analysis_data['summary']['total_results']
        
        for grade in analysis_data['grade_distribution']:
            percentage = (grade['count'] / total_results * 100) if total_results > 0 else 0
            grade_data.append([
                grade['grade'],
                str(grade['count']),
                f"{percentage:.1f}%"
            ])
        
        grade_table = Table(grade_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch])
        grade_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db'))
        ]))
        story.append(grade_table)
        story.append(Spacer(1, 20))
        
        # Subject Performance
        story.append(Paragraph("Subject Performance", heading_style))
        
        subject_data = [['Subject', 'Avg Score', 'Students', 'Pass Rate', 'High Score']]
        
        for subject in analysis_data['subject_performance']:
            pass_rate = (subject['pass_count'] / subject['total_students'] * 100) if subject['total_students'] > 0 else 0
            subject_data.append([
                subject['subject__name'],
                f"{subject['avg_score']:.1f}%",
                str(subject['total_students']),
                f"{pass_rate:.1f}%",
                f"{subject['max_score']:.1f}%"
            ])
        
        subject_table = Table(subject_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        subject_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db'))
        ]))
        story.append(subject_table)
        story.append(Spacer(1, 20))
        
        # Class Performance
        story.append(Paragraph("Class Performance", heading_style))
        
        class_data = [['Class', 'Avg Score', 'Students', 'Pass Rate']]
        
        for class_perf in analysis_data['class_performance']:
            class_data.append([
                class_perf['class_level__name'],
                f"{class_perf['avg_score']:.1f}%",
                str(class_perf['total_students']),
                f"{class_perf['pass_rate']:.1f}%"
            ])
        
        class_table = Table(class_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        class_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db'))
        ]))
        story.append(class_table)
        story.append(Spacer(1, 20))
        
        # Top Performers
        story.append(Paragraph("Top 10 Performers", heading_style))
        
        if analysis_data['top_performers']:
            top_data = [['Student', 'Student ID', 'Subject', 'Class', 'Score', 'Grade']]
            
            for performer in analysis_data['top_performers']:
                top_data.append([
                    f"{performer['student__first_name']} {performer['student__last_name']}",
                    performer['student__student_profile__student_id'],
                    performer['subject__name'],
                    performer['class_level__name'],
                    f"{performer['score']:.1f}%",
                    performer['grade']
                ])
            
            top_table = Table(top_data, colWidths=[1.5*inch, 1*inch, 1.2*inch, 1*inch, 0.8*inch, 0.8*inch])
            top_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db'))
            ]))
            story.append(top_table)
        else:
            story.append(Paragraph("No top performers data available.", styles['Normal']))
        
        # Build PDF
        doc.build(story)
        
        # Get PDF value from buffer
        pdf = buffer.getvalue()
        buffer.close()
        
        # Create HTTP response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="results_analysis_report.pdf"'
        response.write(pdf)
        
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)


def export_excel_report(analysis_data, request):
    """Generate Excel report"""
    try:
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Summary Sheet
        summary_sheet = wb.create_sheet("Executive Summary")
        
        # Header
        summary_sheet.merge_cells('A1:D1')
        summary_sheet['A1'] = "Academic Results Analysis Report"
        summary_sheet['A1'].font = Font(size=16, bold=True, color="1e293b")
        summary_sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Filters info
        filters_text = "Filters: "
        filters = []
        if analysis_data['filters']['academic_year']:
            filters.append(f"Academic Year: {analysis_data['filters']['academic_year'].name}")
        if analysis_data['filters']['class_level']:
            filters.append(f"Class: {analysis_data['filters']['class_level'].name}")
        if analysis_data['filters']['term']:
            filters.append(f"Term: {analysis_data['filters']['term'].name}")
        
        filters_text += " | ".join(filters) if filters else "All Data"
        summary_sheet['A3'] = filters_text
        summary_sheet['A4'] = f"Generated: {analysis_data['generated_at'].strftime('%Y-%m-%d %H:%M')}"
        
        # Summary Table
        summary_sheet['A6'] = "Metric"
        summary_sheet['B6'] = "Value"
        
        summary_data = [
            ['Total Results', analysis_data['summary']['total_results']],
            ['Total Students', analysis_data['summary']['total_students']],
            ['Average Score', analysis_data['summary']['average_score']],
            ['Published Results', analysis_data['summary']['published_results']],
        ]
        
        for i, (metric, value) in enumerate(summary_data, start=7):
            summary_sheet[f'A{i}'] = metric
            summary_sheet[f'B{i}'] = value
        
        # Style summary table
        for row in summary_sheet['A6:B10']:
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
        
        # Grade Distribution Sheet
        grade_sheet = wb.create_sheet("Grade Distribution")
        
        grade_sheet['A1'] = "Grade"
        grade_sheet['B1'] = "Count"
        grade_sheet['C1'] = "Percentage"
        
        for i, grade in enumerate(analysis_data['grade_distribution'], start=2):
            percentage = (grade['count'] / analysis_data['summary']['total_results'] * 100) if analysis_data['summary']['total_results'] > 0 else 0
            grade_sheet[f'A{i}'] = grade['grade']
            grade_sheet[f'B{i}'] = grade['count']
            grade_sheet[f'C{i}'] = percentage / 100 
        
        # Style grade sheet
        for row in grade_sheet['A1:C1']:
            for cell in row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        
        for cell in grade_sheet['C']:
            if cell.row > 1:
                cell.number_format = '0.0%'
        
        subject_sheet = wb.create_sheet("Subject Performance")
        
        subject_headers = ['Subject', 'Code', 'Avg Score', 'Students', 'Pass Count', 'Pass Rate', 'High Score', 'Low Score']
        for col, header in enumerate(subject_headers, start=1):
            subject_sheet.cell(row=1, column=col, value=header)
        
        for i, subject in enumerate(analysis_data['subject_performance'], start=2):
            pass_rate = (subject['pass_count'] / subject['total_students']) if subject['total_students'] > 0 else 0
            subject_sheet.cell(row=i, column=1, value=subject['subject__name'])
            subject_sheet.cell(row=i, column=2, value=subject['subject__code'])
            subject_sheet.cell(row=i, column=3, value=subject['avg_score'])
            subject_sheet.cell(row=i, column=4, value=subject['total_students'])
            subject_sheet.cell(row=i, column=5, value=subject['pass_count'])
            subject_sheet.cell(row=i, column=6, value=pass_rate)
            subject_sheet.cell(row=i, column=7, value=subject['max_score'])
            subject_sheet.cell(row=i, column=8, value=subject.get('min_score', 0))
        
        for col in range(1, len(subject_headers) + 1):
            cell = subject_sheet.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
        
        for row in range(2, len(analysis_data['subject_performance']) + 2):
            subject_sheet.cell(row=row, column=6).number_format = '0.0%'
            subject_sheet.cell(row=row, column=3).number_format = '0.0'
            subject_sheet.cell(row=row, column=7).number_format = '0.0'
            subject_sheet.cell(row=row, column=8).number_format = '0.0'
        
        class_sheet = wb.create_sheet("Class Performance")
        
        class_headers = ['Class', 'Avg Score', 'Students', 'Pass Rate']
        for col, header in enumerate(class_headers, start=1):
            class_sheet.cell(row=1, column=col, value=header)
        
        for i, class_perf in enumerate(analysis_data['class_performance'], start=2):
            class_sheet.cell(row=i, column=1, value=class_perf['class_level__name'])
            class_sheet.cell(row=i, column=2, value=class_perf['avg_score'])
            class_sheet.cell(row=i, column=3, value=class_perf['total_students'])
            class_sheet.cell(row=i, column=4, value=class_perf['pass_rate'] / 100)
        
        for col in range(1, len(class_headers) + 1):
            cell = class_sheet.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="8b5cf6", end_color="8b5cf6", fill_type="solid")
        
        for row in range(2, len(analysis_data['class_performance']) + 2):
            class_sheet.cell(row=row, column=2).number_format = '0.0'
            class_sheet.cell(row=row, column=4).number_format = '0.0%'
        
        if analysis_data['top_performers']:
            top_sheet = wb.create_sheet("Top Performers")
            
            top_headers = ['Rank', 'Student Name', 'Student ID', 'Subject', 'Class', 'Score', 'Grade']
            for col, header in enumerate(top_headers, start=1):
                top_sheet.cell(row=1, column=col, value=header)
            
            for i, performer in enumerate(analysis_data['top_performers'], start=2):
                top_sheet.cell(row=i, column=1, value=i-1)
                top_sheet.cell(row=i, column=2, value=f"{performer['student__first_name']} {performer['student__last_name']}")
                top_sheet.cell(row=i, column=3, value=performer['student__student_profile__student_id'])
                top_sheet.cell(row=i, column=4, value=performer['subject__name'])
                top_sheet.cell(row=i, column=5, value=performer['class_level__name'])
                top_sheet.cell(row=i, column=6, value=performer['score'])
                top_sheet.cell(row=i, column=7, value=performer['grade'])
            
            for col in range(1, len(top_headers) + 1):
                cell = top_sheet.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid")
            
            for row in range(2, len(analysis_data['top_performers']) + 2):
                top_sheet.cell(row=row, column=6).number_format = '0.0'
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for col in ws.columns:
                max_length = 0
                col_letter = None
                
                for cell in col:
                    if isinstance(cell, MergedCell):
                        continue
                    
                    if col_letter is None:
                        col_letter = cell.column_letter
                    
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                if col_letter:
                    ws.column_dimensions[col_letter].width = (max_length + 2) * 1.2
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="results_analysis_report.xlsx"'
        
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating Excel report: {str(e)}", status=500)
    

@login_required
def terms_list(request):
    """
    Main terms list view - renders the template
    """
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    
    context = {
        'academic_years': academic_years,
    }
    
    return render(request, "pages/academics/terms.html", context)


@login_required
def api_terms_list(request):
    try:
        academic_year_id = request.GET.get('academic_year') or None
        status_filter = request.GET.get('status')
        search_query = request.GET.get('search', '')
        page_number = request.GET.get('page', 1)

        # Fix page_number errors
        try:
            page_number = int(page_number)
        except:
            page_number = 1
        
        terms = Term.objects.select_related('academic_year').all()

        # Safe check for academic_year
        if academic_year_id not in [None, ""] and academic_year_id.isdigit():
            terms = terms.filter(academic_year_id=int(academic_year_id))


        # Status filters
        if status_filter:
            today = timezone.now().date()
            if status_filter == 'current':
                terms = terms.filter(is_current=True)
            elif status_filter == 'upcoming':
                terms = terms.filter(start_date__gt=today)
            elif status_filter == 'past':
                terms = terms.filter(end_date__lt=today)
            elif status_filter == 'active':
                terms = terms.filter(start_date__lte=today, end_date__gte=today)

        # Search filter
        if search_query:
            terms = terms.filter(
                Q(name__icontains=search_query) |
                Q(academic_year__name__icontains=search_query)
            )

        terms = terms.order_by('-start_date')

        paginator = Paginator(terms, 10)
        page_obj = paginator.get_page(page_number)

        terms_data = []
        for term in page_obj:
            terms_data.append({
                'id': term.id,
                'name': term.name,
                'academic_year_id': term.academic_year.id if term.academic_year else None,
                'academic_year_name': term.academic_year.name if term.academic_year else None,
                'start_date': term.start_date.isoformat(),
                'end_date': term.end_date.isoformat(),
                'vacation_start_date': term.vacation_start_date.isoformat() if term.vacation_start_date else None,
                'vacation_end_date': term.vacation_end_date.isoformat() if term.vacation_end_date else None,
                'half_term_start_date': term.half_term_start_date.isoformat() if term.half_term_start_date else None,
                'half_term_end_date': term.half_term_end_date.isoformat() if term.half_term_end_date else None,
                'holidays': term.holidays or '',
                'school_activities': term.school_activities or '',
                'is_current': term.is_current,
            })

        pagination_data = {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_count': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'start_index': page_obj.start_index(),
            'end_index': page_obj.end_index(),
        }

        return JsonResponse({
            'success': True,
            'terms': terms_data,
            'pagination': pagination_data
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)  # show the error for debugging
        }, status=500)


@login_required
@require_http_methods(["POST"])
def api_update_term(request):
    """
    API endpoint to update a term
    """
    try:
        term_id = request.POST.get('term_id')
        term = get_object_or_404(Term, id=term_id)
        
        name = request.POST.get('name')
        academic_year_id = request.POST.get('academic_year')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        is_current = request.POST.get('is_current') == 'on'
        
        if not all([name, academic_year_id, start_date, end_date]):
            return JsonResponse({
                'success': False,
                'error': 'All fields are required'
            }, status=400)
        
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        
        if Term.objects.filter(
            name=name, 
            academic_year=academic_year
        ).exclude(id=term_id).exists():
            return JsonResponse({
                'success': False,
                'error': f'A term with name "{name}" already exists in {academic_year.name}'
            }, status=400)
        
        term.name = name
        term.academic_year = academic_year
        term.start_date = start_date
        term.end_date = end_date
        
        if is_current:
            set_current_term(term)
        else:
            term.is_current = False
            term.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Term "{term.name}" updated successfully'
        })
        
    except Exception as e:
        # logger.error(f"Error updating term: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Failed to update term'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def api_delete_term(request):
    """
    API endpoint to delete a term
    """
    try:
        data = json.loads(request.body)
        term_id = data.get('term_id')
        term = get_object_or_404(Term, id=term_id)
        
        # Check if term has related records
        if term.results.exists():
            return JsonResponse({
                'success': False,
                'error': 'Cannot delete term with associated results'
            }, status=400)
        
        term_name = term.name
        term.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Term "{term_name}" deleted successfully'
        })
        
    except Exception as e:
        # logger.error(f"Error deleting term: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Failed to delete term'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def api_set_current_term(request):
    """
    API endpoint to set a term as current
    """
    try:
        data = json.loads(request.body)
        term_id = data.get('term_id')
        term = get_object_or_404(Term, id=term_id)
        
        set_current_term(term)
        
        return JsonResponse({
            'success': True,
            'message': f'Term "{term.name}" set as current term'
        })
        
    except Exception as e:
        # logger.error(f"Error setting current term: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Failed to set current term'
        }, status=500)
    

@login_required
@require_http_methods(["GET", "POST"])
def create_term(request):
    """
    Handle term creation - both form display and form submission
    """
    if request.method == 'GET':
        return render_create_term_form(request)
    else:
        return handle_term_creation(request)


@login_required
def render_create_term_form(request):
    """
        Render the term creation form with required context
    """
    academic_years = AcademicYear.objects.filter(
        end_date__gte=timezone.now().date()
    ).order_by('-start_date')
    
    context = {
        'academic_years': academic_years,
    }
    
    return render(request, 'pages/academics/create_term.html', context)


@transaction.atomic
def handle_term_creation(request):
    """
        Process term creation form submission
    """
    try:
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        term_data = extract_term_data(request)
        
        validation_errors = validate_term_data(term_data)
        if validation_errors:
            return error_response(validation_errors, is_ajax)
        
        academic_year = get_object_or_404(
            AcademicYear, 
            id=term_data['academic_year_id']
        )
        
        if Term.objects.filter(
            name=term_data['name'],
            academic_year=academic_year
        ).exists():
            error_msg = f"A term with name '{term_data['name']}' already exists in {academic_year.name}."
            return error_response(error_msg, is_ajax)
        
        date_errors = validate_date_ranges(term_data, academic_year)
        if date_errors:
            return error_response(date_errors, is_ajax)
        
        overlap_errors = check_term_overlaps(term_data, academic_year)
        if overlap_errors:
            return error_response(overlap_errors, is_ajax)
        
        
        term = create_term_instance(term_data, academic_year, request.user)
        
        if term_data.get('is_current'):
            set_current_term(term)
        
        success_message = f"Term '{term.name}' for {academic_year.name} created successfully!"
        
        if is_ajax:
            return JsonResponse({
                'success': True,
                'message': success_message,
                'term_id': term.id,
            })
        else:
            messages.success(request, success_message)
            return redirect('terms_list')
            
    except Exception as e:
        # logger.error(f"Error creating term: {str(e)}")
        error_msg = "An unexpected error occurred while creating the term."
        
        if is_ajax:
            return JsonResponse({
                'success': False,
                'error': error_msg
            }, status=500)
        else:
            messages.error(request, error_msg)
            return redirect('create_term')


def extract_term_data(request):
    """
    Extract and clean term data from request
    """
    if request.content_type == 'application/json':
        data = json.loads(request.body)
    else:
        data = request.POST
    
    return {
        'name': data.get('name', '').strip(),
        'academic_year_id': data.get('academic_year'),
        'start_date': data.get('start_date'),
        'end_date': data.get('end_date'),
        'is_current': data.get('is_current') == 'true' or data.get('is_current') == 'on',
    }


def validate_term_data(term_data):
    """
    Validate required term data
    """
    errors = []
    
    if not term_data['name']:
        errors.append("Term name is required.")
    
    if not term_data['academic_year_id']:
        errors.append("Academic year is required.")
    
    if not term_data['start_date']:
        errors.append("Start date is required.")
    
    if not term_data['end_date']:
        errors.append("End date is required.")
    
    valid_terms = ['1st Term', '2nd Term', '3rd Term']
    if term_data['name'] and term_data['name'] not in valid_terms:
        errors.append(f"Term name must be one of: {', '.join(valid_terms)}")
    
    return errors


def validate_date_ranges(term_data, academic_year):
    """
    Validate date ranges and relationships
    """
    errors = []
    
    try:
        start_date = datetime.strptime(term_data['start_date'], '%Y-%m-%d').date()
        end_date = datetime.strptime(term_data['end_date'], '%Y-%m-%d').date()
        
        # Checking if start date is before end date
        if start_date >= end_date:
            errors.append("End date must be after start date.")
        
        # Checking if term dates are within academic year
        if start_date < academic_year.start_date:
            errors.append(f"Term start date cannot be before academic year start date ({academic_year.start_date}).")
        
        if end_date > academic_year.end_date:
            errors.append(f"Term end date cannot be after academic year end date ({academic_year.end_date}).")
        
        # Checking if term duration is reasonable (at least 1 week)
        term_duration = (end_date - start_date).days
        if term_duration < 7:
            errors.append("Term duration must be at least 1 week.")
            
        if term_duration > 365:
            errors.append("Term duration cannot exceed 1 year.")
            
    except ValueError:
        errors.append("Invalid date format. Please use YYYY-MM-DD format.")
    
    return errors


def check_term_overlaps(term_data, academic_year):
    """
    Check for overlapping terms in the same academic year
    """
    errors = []
    
    try:
        start_date = datetime.strptime(term_data['start_date'], '%Y-%m-%d').date()
        end_date = datetime.strptime(term_data['end_date'], '%Y-%m-%d').date()
        
        overlapping_terms = Term.objects.filter(
            academic_year=academic_year
        ).exclude(
            name=term_data['name'] 
        ).filter(
            start_date__lte=end_date,
            end_date__gte=start_date
        )
        
        if overlapping_terms.exists():
            overlapping_names = [term.name for term in overlapping_terms]
            errors.append(
                f"Term dates overlap with existing term(s): {', '.join(overlapping_names)}. "
                f"Please adjust the dates."
            )
            
    except ValueError:
        pass
    
    return errors


def create_term_instance(term_data, academic_year, user):
    """
    Create and save the term instance
    """
    term = Term(
        name=term_data['name'],
        academic_year=academic_year,
        start_date=term_data['start_date'],
        end_date=term_data['end_date'],
        is_current=False
    )
    
    term.save()
    return term


def set_current_term(new_current_term):
    """
    Set a term as current and unset any previous current term
    """
    Term.objects.filter(
        academic_year=new_current_term.academic_year,
        is_current=True
    ).update(is_current=False)
    
    new_current_term.is_current = True
    new_current_term.save()



def error_response(request, error_message, is_ajax):
    """
    Return appropriate error response based on request type
    """
    if is_ajax:
        return JsonResponse({
            'success': False,
            'error': error_message if isinstance(error_message, str) else ' '.join(error_message)
        }, status=400)
    else:
        if isinstance(error_message, list):
            error_message = ' '.join(error_message)
        messages.error(request, error_message)
        return redirect('create_term')


@login_required
def get_terms_by_academic_year(request):
    """
    API endpoint to get terms for a specific academic year
    """
    academic_year_id = request.GET.get('academic_year_id')
    
    if not academic_year_id:
        return JsonResponse({
            'success': False,
            'error': 'Academic year ID is required'
        }, status=400)
    
    try:
        terms = Term.objects.filter(academic_year_id=academic_year_id).order_by('start_date')
        
        terms_data = []
        for term in terms:
            terms_data.append({
                'id': term.id,
                'name': term.name,
                'start_date': term.start_date.strftime('%Y-%m-%d'),
                'end_date': term.end_date.strftime('%Y-%m-%d'),
                'vacation_start_date': term.vacation_start_date.strftime('%Y-%m-%d') if term.vacation_start_date else '',
                'vacation_end_date': term.vacation_end_date.strftime('%Y-%m-%d') if term.vacation_end_date else '',
                'half_term_start_date': term.half_term_start_date.strftime('%Y-%m-%d') if term.half_term_start_date else '',
                'half_term_end_date': term.half_term_end_date.strftime('%Y-%m-%d') if term.half_term_end_date else '',
                'holidays': term.holidays or '',
                'school_activities': term.school_activities or '',
                'is_current': term.is_current,
                'academic_year_name': term.academic_year.name
            })
        
        return JsonResponse({
            'success': True,
            'terms': terms_data
        })
        
    except AcademicYear.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Academic year not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
